""" This module is used to upgrade the previous version of excel file from 2.0 to the latest version (V3.0). """
import argparse
import json
import os
import warnings

from openpyxl import load_workbook
from rich.prompt import Confirm

from base.namespace import Language
from base.source.infosheet import InfoNode, Sheet, Sheets
from base.source.tablesheet import Table, TableNode, Tables
from base.source.writer import ExcelWritter
from base.utils.client.utils import get_immenv_value
from base.utils.client.utils import append_ext
from client.system.config import BASEDIR
import getpass

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class OldExcel:
    """Excel manupulate excel for getting source data, compare excels or sheets data, and write new excel for collecting data.

    Args:
        excel_name ([str]): [an excel file name]
        sheets ([list]): [default is None, meaning will get all sheets; If giving a list of sheet names, it will only access the specified sheets.
        sheet name in this Excel class will with its original 'info-' or 'table-' head, but the head will be removed when save to dict for exporting;
        When write to a new excel file, the sheet name of a object of Sheet will be insert 'info-' in the name, while for Table object, 'table-'
        will be inserted too]
    Returns: an Excel object
    """

    def __init__(self, excel_name, language=Language.ENGLISH):
        self.language = language
        self.excel_name = excel_name
        self.schema = self.get_the_schema()
        self._read_excel()

    def get_the_schema(self):
        with open(BASEDIR / "base/source/schema.json", "r") as f:
            schema = json.load(f)
        return schema

    def _read_excel(self):
        self.wb = load_workbook(self.excel_name)
        """ if not specify sheets, it will handle all sheets """
        self.sheetNames = self.wb.sheetnames
        self.sheet_names = [
            sheetName
            for sheetName in self.sheetNames
            if sheetName.lower().startswith("info-")
        ]  # save info sheet names
        self.table_names = [
            sheetName
            for sheetName in self.sheetNames
            if sheetName.lower().startswith("table-")
        ]  # save table sheet names
        """ self._sheets and self._tables are representing the whole excel data. All operations will be done on these two variables."""
        self._sheets = None  # save info sheet objects
        self._tables = None  # save table sheet objects
        self._read_sheets_tables()  # get all info and table sheets objects and save to sheets and tables

        """
        Below are methods for getting data from excel files and export 
        """

    """ get the index based on the variable name and sheet/table name"""

    def get_index(self, variable_name, sheet_table_name):
        table_dict = self.schema.get(sheet_table_name)
        if table_dict and table_dict.get(variable_name):
            return table_dict.get(variable_name).get("index")

    # get the excel's specified sheets(info and table). In this step, the 'info-' and 'table-' will be removed. The rest sheet name
    # will be used as the key of the object of Sheet or Table.
    def _read_sheets_tables(self):
        sheets = [self._read_sheet(sn) for sn in self.sheet_names]
        self._sheets = Sheets(sheets=sheets)
        talbes = [self._read_table(sn) for sn in self.table_names]
        self._tables = Tables(tables=talbes)

    def _read_sheet(self, sheet):
        values = list(self.wb[sheet].values)
        """ clean data by removing None """
        values = [x for x in values if self._notNone(x)]
        sheet_title = values[0][0]
        column_titiles = list(values[1])
        column_titiles[1] = "Index"  # upgrade the index column title
        column_variable = [str(var) for var in values[2]]
        column_variable[1] = "index"  # upgrade the index column variable
        data = []
        for data_row in values[3:]:
            data_row = list(data_row)
            the_index = self.get_index(data_row[0], sheet)
            # if the_index == None, means new sheet has deleted some variables, so the index will be None. So we just ignore the data
            if the_index == None:
                continue
            data_row[1] = the_index
            texts = [text for text in data_row]
            info_node = InfoNode(**dict(zip(column_variable, texts)))
            data.append(info_node)

        sheet_obj = Sheet(
            sheet_name=sheet,
            sheet_title=sheet_title,
            column_titles=column_titiles,
            column_variables=column_variable,
            data=data,
        )
        return sheet_obj

    # read table sheet data from excel
    def _read_table(self, table):
        values = list(self.wb[table].values)
        """ clean data by removing None """
        values = [x for x in values if self._notNone(x)]
        table_title = values[0][0]
        column_titles = values[1]
        column_variable = values[2]

        # check if row 4 has index alreay, if not, add it
        is_version_3 = False
        if len(values) > 3:
            value_list_is_number = [
                True if type(value) == int and value % 5 == 0 else False
                for value in values[3]
            ]
            is_version_3 = True if all(value_list_is_number) else False
        if is_version_3:
            variable_index = values[3]
        else:
            variable_index = [
                self.get_index(variable, table) for variable in column_variable
            ]
            # if some index inside the list are None, means new sheet has deleted some variables, So we set the index to 99999
            variable_index = [i if i else 99999 for i in variable_index]

        data_starts = 4 if is_version_3 else 3
        data = []
        for data_row in values[data_starts:]:
            texts = [text for i, text in enumerate(data_row)]
            table_node = TableNode(**dict(zip(column_variable, texts)))
            data.append(table_node)

        new_table_obj = Table(
            table_name=table,
            table_title=table_title,
            column_titles=column_titles,
            column_variables=column_variable,
            column_index=variable_index,
            data=data,
        )
        return new_table_obj

    """ check if a row is None. Return True if there is at least one column not None """

    def _notNone(self, row_data):
        return any([False if x == None else True for x in row_data])


def upgrade(
    file,
    output_file_name,
    password,
    language=Language.ENGLISH,
    protection=True,
    ignore_overwrite=False,
):
    _, ext = os.path.splitext(file)
    if ext != ".xlsx":
        return
    input_file_is_existed = os.path.exists(file)
    if not input_file_is_existed:
        print(f"{file} is not existed")
        exit(1)
    output_file_is_existed = os.path.exists(output_file_name)
    if (
        not ignore_overwrite
        and output_file_is_existed
        and not Confirm.ask(
            f"{output_file_name} existed. Do you want to overwrite it? "
        )
    ):
        return

    excel = OldExcel(file)
    output = ExcelWritter(
        output_file_name=output_file_name,
        sheets=excel._sheets._sheets,
        tables=excel._tables._tables,
        language=language,
    )
    output.create(protection=protection, password=password)
    print(f"excel file {output_file_name} has been upgraded to V3.0")


# get only excel files
def get_excel_files(directory):
    files = os.listdir(directory)
    excel_files = []
    for file in files:
        _, ext = os.path.splitext(file)
        if ext == ".xlsx":
            excel_files.append(file)
    return excel_files


def main():
    parser = argparse.ArgumentParser(
        "This is a tool for upgrading excel file to V3.0. The source could be file or directory, while the target is same, but you have to keep them syncronized ."
    )
    parser.add_argument("source", type=str, help="Source excel file or directory")
    parser.add_argument(
        "target", type=str, help="Target excel file or directory", nargs="?"
    )
    parser.add_argument("-c", "--chinese", help="Use Chinese", action="store_true")
    parser.add_argument(
        "-po", "--protection_off", help="Do not use protection", action="store_true"
    )
    parser.add_argument(
        "-y",
        "--yes_to_all",
        help="Yes to all, and ignore the warning and overwite existed files automatically",
        action="store_true",
    )

    args = parser.parse_args()

    language = Language.CHINESE if args.chinese else Language.ENGLISH
    password = (
        get_immenv_value("EXCEL_PASSWORD") or getpass.getpass("Password: ")
        if not args.protection_off
        else ""
    )

    if os.path.isdir(args.source):
        """source is a directory"""
        excel_files = get_excel_files(args.source)
        target_directory = args.target or args.source
        for excel_file in excel_files:
            upgrade(
                f"{args.source}/{excel_file}",
                f"{target_directory}/{excel_file}",
                password,
                language,
                protection=not args.protection_off,
                ignore_overwrite=args.yes_to_all,
            )
    elif os.path.isfile(append_ext(args.source, ".xlsx")):
        """source is a file"""
        source = append_ext(args.source, ".xlsx")
        target = append_ext(args.target or args.source, ".xlsx")
        upgrade(
            source,
            target,
            password,
            language,
            protection=not args.protection_off,
            ignore_overwrite=args.yes_to_all,
        )

    else:
        print(f"{args.source} does not exist.")
        exit(1)


if __name__ == "__main__":
    main()
