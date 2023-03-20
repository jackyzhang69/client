"""
This module is used to write the data to excel file with given format defined in base/source1/formatter.py
"""


import xlsxwriter
from openpyxl.utils import get_column_letter
from base.namespace import Language
from typing import Optional
from base.source.infosheet import Sheets
from base.source.tablesheet import Tables
from dataclasses import dataclass, field
from base.source.formatter import *
from base.utils.client.utils import Toggle
from typing import Union


@dataclass
class NodeProperties:
    language: Language = Language.ENGLISH
    comment: str = ""
    validation: Optional[dict] = field(default_factory=dict)
    style: Optional[dict] = field(default_factory=dict)
    width: Optional[int] = 10  # default width of table column
    other: Optional[dict] = field(default_factory=dict)

    """ Get properties of a variable by checking the cell_formatter definition"""

    def get_var_properties(self, sheet_table_name, variable):
        cell_formatter = CellFormatter(self.language).cell_formatter
        if sheet_table_name in cell_formatter.keys():
            if variable in cell_formatter[sheet_table_name].keys():
                formatter = cell_formatter[sheet_table_name][variable]
                self.comment = (
                    formatter["comment"][self.language.value]
                    if "comment" in formatter.keys()
                    else ""
                )
                self.validation = (
                    formatter["validation"] if "validation" in formatter.keys() else {}
                )
                self.style = formatter["style"] if "style" in formatter.keys() else {}
                self.width = formatter["width"] if "width" in formatter.keys() else None
                self.other = formatter["other"] if "other" in formatter.keys() else {}
                return self
            else:
                return None


class ExcelWritter:
    def __init__(
        self,
        output_file_name: str,
        sheets: Union[Sheets, None] = None,
        tables: Union[Tables, None] = None,
        language=Language.ENGLISH,
    ):
        self.workbook = xlsxwriter.Workbook(output_file_name)
        self.sheets = sheets
        self.tables = tables
        self.language = language
        self.workbook.set_size(1440, 1640)
        self.protection_options = PROTECTION_OPTIONS
        self.setup_default_formats()

    def setup_default_formats(self):
        self.title_format = self.workbook.add_format(TITLE_FORMAT)
        self.column_title_format = self.workbook.add_format(COLUMN_TITLE_FORMAT)
        self.variable_title_format = self.workbook.add_format(VARIABLE_TITLE_FORMAT)
        self.description_format = self.workbook.add_format(DESCRIPTION_FORMAT)
        self.description_format.set_text_wrap()
        self.money = self.workbook.add_format({"num_format": "$#,##0"})
        self.date_format = self.workbook.add_format({"num_format": "yyyy-mm-dd"})
        self.integer_format = self.workbook.add_format({"num_format": "#,##0"})
        self.float_format = self.workbook.add_format({"num_format": "0.00"})
        self.percent_format0 = self.workbook.add_format({"num_format": "0.%"})
        self.percent_format1 = self.workbook.add_format({"num_format": "0.0%"})
        self.percent_format2 = self.workbook.add_format({"num_format": "0.00%"})
        self.default_format = self.workbook.add_format({"num_format": "@"})
        self.default_format.set_text_wrap()
        self.comment_format = COMMENT_FORMAT

    def get_sheet_table_title_note(self, sheet_table_name):
        cell_formatter = CellFormatter(self.language).cell_formatter
        if sheet_table_name in cell_formatter.keys():
            if "title_note" in cell_formatter[sheet_table_name].keys():
                title_note = cell_formatter[sheet_table_name]["title_note"][
                    self.language.value
                ]
                return title_note

    """ Grouping numbers in a list into a dict, the key is the int part of the number, and the value is a list of numbers"""

    def group_numbers(self, numbers):
        groups = {}
        for number in numbers:
            int_part = int(number)
            if int_part in groups:
                groups[int_part].append(number)
            else:
                groups[int_part] = [number]

        groups = {key: value for key, value in groups.items() if len(value) > 1}
        return groups

    """ Loop the sheets and write each sheet """

    def write_sheets(self, protection, password):
        if self.sheets == None or self.sheets == []:
            return None
        for sheet_content in self.sheets:
            sheet = self.workbook.add_worksheet(sheet_content.sheet_name)
            self.write_sheet(sheet, sheet_content, protection, password)

    """ Loop the tables and write each table """

    def write_tables(self, protection, password):
        if self.tables == None or self.tables == []:
            return None

        for table_content in self.tables:
            table = self.workbook.add_worksheet(table_content.table_name)
            self.write_table(table, table_content, protection, password)

    """Write the sheet header"""

    def write_sheet_header(self, sheet, sheet_content):
        """Write the sheet title"""
        sheet.merge_range(0, 0, 0, 3, None, self.title_format)
        sheet.write(0, 0, sheet_content.sheet_title, self.title_format)

        title_note = self.get_sheet_table_title_note(sheet_content.sheet_name)
        if title_note:
            sheet.write_comment(0, 0, title_note, self.comment_format)

        """ Write the column titls"""
        for index, column_title in enumerate(sheet_content.column_titles):
            sheet.write(1, index, column_title, self.column_title_format)

        """ Write the column variables"""
        for index, column_variable in enumerate(sheet_content.column_variables):
            sheet.write(2, index, column_variable, self.variable_title_format)

    """Write the table header"""

    def write_table_header(self, table, table_content):
        """Merge title cell"""
        columns = len(table_content.column_titles)
        table.merge_range(0, 0, 0, columns - 1, None, self.title_format)

        """ Write the table title"""
        table.write(0, 0, table_content.table_title, self.title_format)
        title_note = self.get_sheet_table_title_note(table_content.table_name)
        if title_note:
            table.write_comment(0, 0, title_note, self.comment_format)

        """ Write the column titles"""
        for index, column_title in enumerate(table_content.column_titles):
            table.write(1, index, column_title, self.column_title_format)

            variable = table_content.column_variables[index]
            table_node_properties = NodeProperties(
                language=self.language
            ).get_var_properties(table_content.table_name, variable)

            if table_node_properties and table_node_properties.comment:
                table.write_comment(
                    1, index, table_node_properties.comment, self.comment_format
                )

        """ Write the column variables"""
        for index, column_variable in enumerate(table_content.column_variables):
            table.write(2, index, column_variable, self.variable_title_format)

        """ Write the column index"""
        for index, column_index in enumerate(table_content.column_index):
            table.write(3, index, column_index, self.variable_title_format)

    """ Get the value format"""

    def get_value_format(self, sheet_table_content_name, variable):
        table_node_properties = NodeProperties(
            language=self.language
        ).get_var_properties(sheet_table_content_name, variable)

        value_format_dict = VALUE_FORMAT
        if table_node_properties and table_node_properties.style:
            value_format_dict = {**value_format_dict, **table_node_properties.style}
        value_format = self.workbook.add_format(value_format_dict)
        return table_node_properties, value_format

    """ Format sheet"""

    def get_sheet_width(self, worksheet):
        # get description and value width from formatter
        cell_formatter = CellFormatter(self.language).cell_formatter
        description_width = (
            cell_formatter[worksheet.name]["description_width"]
            if cell_formatter.get(worksheet.name)
            and cell_formatter.get(worksheet.name).get("description_width")
            else None
        )
        value_width = (
            cell_formatter[worksheet.name]["value_width"]
            if cell_formatter.get(worksheet.name)
            and cell_formatter.get(worksheet.name).get("value_width")
            else None
        )
        return description_width, value_width

    def format_sheet(self, worksheet, protection, password):
        """and set column description and value width"""
        description_width, value_width = self.get_sheet_width(worksheet)
        desciption_width = description_width or SHEET_DESCRIPTION_WIDTH
        value_width = value_width or SHEET_VALUE_WIDTH

        worksheet.set_column("A:B", 15, None, {"hidden": True})
        worksheet.set_column("C:C", desciption_width, None)
        worksheet.set_column("D:D", value_width, None)

        """ Hide columns variable and index, , and hide variable row"""
        worksheet.set_column("A:B", 15, None, {"hidden": True})
        worksheet.set_row(2, None, None, {"hidden": True})

        # protect the sheet
        if protection:
            worksheet.protect(password, self.protection_options)

    """ Format table"""

    def format_table(self, worksheet, table_content, protection, password):

        """Set column width based on formatter"""
        for col_index, column_variable in enumerate(table_content.column_variables):
            # get column width from formatter
            var_property = NodeProperties(language=self.language).get_var_properties(
                table_content.table_name, column_variable
            )
            width = (
                var_property.width
                if var_property and var_property.width
                else TABLE_VALUE_WIDTH
            )
            column = self.get_col_symbol(col_index)
            worksheet.set_column(column, width, self.default_format)

        """Hide rows variable and index"""
        worksheet.set_row(2, None, None, {"hidden": True})
        worksheet.set_row(3, None, None, {"hidden": True})
        
        """ Hide variable column"""
        if table_content.table_name in SPECIAL_TABLE_LIST:
            worksheet.set_column("A:A", 15, None, {"hidden": True})

        # protect the sheet
        if protection:
            worksheet.protect(password, self.protection_options)

    """ Get column symbol by column index, return strig as exp: "A1" """

    def get_col_symbol(self, col_index):
        return get_column_letter(col_index + 1) + ":" + get_column_letter(col_index + 1)

    """ Write a sheet """

    # write a row of the sheet
    def write_a_row(
        self,
        sheet,
        row_index,
        row,
        variable,
        info_node_properties,
        description_format,
        value_format,
    ):
        # write variables
        sheet.write(row_index + 3, 0, variable, self.default_format)
        # write index
        sheet.write(row_index + 3, 1, row.index, self.default_format)
        # write description
        sheet.write(row_index + 3, 2, row.description, description_format)
        # write comment
        if info_node_properties and info_node_properties.comment:
            sheet.write_comment(
                row_index + 3, 2, info_node_properties.comment, self.comment_format
            )
        # write value
        sheet.write(row_index + 3, 3, row.value, value_format)

    def write_sheet(self, sheet, sheet_content, protection, password):
        self.write_sheet_header(sheet, sheet_content)

        """ Write the sheet content"""
        # sort the sheet content
        sheet_content = sheet_content.get_sorted_sheet()

        # group the indexes
        indexes = [float(data.index) for data in sheet_content.data]
        groups = self.group_numbers(indexes)

        # assign the groups with toggled color
        group_colors = {}
        for group_index in groups.keys():
            group_colors[group_index] = Toggle().toggle_color()

        for row_index, row in enumerate(sheet_content.data):
            # get row data properties of validation, comment, style, etc.
            variable = row.variable

            info_node_properties, value_format = self.get_value_format(
                sheet_content.sheet_name, variable
            )

            # check if variable belongs to a group, if yes, set the group color
            description_format = self.workbook.add_format(DESCRIPTION_FORMAT)
            variable_index_int_part = int(float(row.index))
            if variable_index_int_part in groups.keys():
                # get the group color
                description_format.set_bg_color(group_colors[variable_index_int_part])
                value_format.set_bg_color(group_colors[variable_index_int_part])

            self.write_a_row(
                sheet,
                row_index,
                row,
                variable,
                info_node_properties,
                description_format,
                value_format,
            )

            # Create a DataValidation object
            if info_node_properties and info_node_properties.validation:
                sheet.data_validation(
                    row_index + 3, 3, row_index + 3, 3, info_node_properties.validation
                )

        self.format_sheet(sheet, protection, password)

    """ Write a table """

    def write_table(self, table, table_content, protection, password):
        self.write_table_header(table, table_content)

        """ Write the table content"""
        # sort the table by index
        table_content = table_content.get_sorted_table()

        """ Format a range of table cells based on the cell properties"""
        for column_index, variable in enumerate(table_content.column_variables):
            table_node_properties, value_format = self.get_value_format(
                table_content.table_name, variable
            )

            """ foramt 100 rows"""
            if table_content.table_name not in SPECIAL_TABLE_LIST:
                for i in range(PRE_FORMAT_ROWS_IN_TABLE):
                    table.write(i + 4, column_index, "", value_format)

                # Create a DataValidation object
                if table_node_properties and table_node_properties.validation:
                    table.data_validation(
                        4,
                        column_index,
                        PRE_FORMAT_ROWS_IN_TABLE + 4,
                        column_index,
                        table_node_properties.validation,
                    )

        """ write the table content """
        for row_index, data in enumerate(table_content.data):
            """Convert TableNode object to list of values"""
            values = [value for value in data.__dict__.values()]
            for column_index, value in enumerate(values):
                variable = table_content.column_variables[column_index]
                _, value_format = self.get_value_format(
                    table_content.table_name, variable
                )

                value = (
                    float(value) if value and "#," in value_format.num_format else value
                )
                table.write(row_index + 4, column_index, value, value_format)

        self.format_table(table, table_content, protection, password)

    def create(self, protection=True, password=""):
        self.write_sheets(protection, password)
        self.write_tables(protection, password)
        self.workbook.close()
