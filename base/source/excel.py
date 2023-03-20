from base.source.infosheet import InfoNode
from base.source.infosheet import Sheet, Sheets
from base.source.tablesheet import TableNode, Table, Tables
from base.source.writer import ExcelWritter
from base.namespace import Language
from openpyxl import load_workbook
from base.source.source import Source
import warnings
import copy


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class Excel(Source):
    """Excel manupulate excel for getting source data, compare excels or sheets data, and write new excel for collecting data.

    Args:
        excel_name (str): an excel file name
        language: an enum of Language
    Returns: an Excel object
    """

    def __init__(self, excel_name, language=Language.ENGLISH):
        self.language = language
        self.excel_name = excel_name
        self._read_excel()

    def _read_excel(self):
        self.wb = load_workbook(self.excel_name)
        self.sheet_names = [
            sheetName
            for sheetName in self.wb.sheetnames
            if sheetName.lower().startswith("info-")
        ]
        self.table_names = [
            sheetName
            for sheetName in self.wb.sheetnames
            if sheetName.lower().startswith("table-")
        ]

        self._read_sheets_tables()  # get all info and table sheets objects and save to sheets and tables

        """
        Below are methods for getting data from excel files and export 
        """

    # will be used as the key of the object of Sheet or Table.
    def _read_sheets_tables(self):
        sheets = [self._read_sheet(sn) for sn in self.sheet_names]
        self._sheets = Sheets(sheets=sheets)
        talbes = [self._read_table(sn) for sn in self.table_names]
        self._tables = Tables(tables=talbes)

    def _read_sheet(self, sheet):
        values = list(self.wb[sheet].values)
        """ clean data by removing None """
        values = [x for x in values if self._notNone(x)]
        sheet_title = values[0][0]
        column_titiles = values[1]
        column_variable = [str(var) for var in values[2]]
        data = []
        for data_row in values[3:]:
            texts = [text if text != None else None for text in data_row]
            info_node = InfoNode(**dict(zip(column_variable, texts)))
            data.append(info_node)

        sheet_obj = Sheet(
            sheet_name=sheet,
            sheet_title=sheet_title,
            column_titles=column_titiles,
            column_variables=column_variable,
            data=data,
        )
        return sheet_obj

    def _read_table(self, table):
        values = list(self.wb[table].values)
        """ clean data by removing None """
        values = [x for x in values if self._notNone(x)]
        table_title = values[0][0]
        column_titles = values[1]
        column_variable = values[2]
        variable_index = [var for var in values[3]]
        data = []
        for data_row in values[4:]:
            texts = [text if text != None else None for text in data_row]
            if not any(texts):
                continue
            table_node = TableNode(**dict(zip(column_variable, texts)))
            data.append(table_node)

        table_obj = Table(
            table_name=table,
            table_title=table_title,
            column_titles=column_titles,
            column_variables=column_variable,
            column_index=variable_index,
            data=data,
        )
        return table_obj

    # get sheet object from the dict of self.sheets by specific sheet name,and variable list. default None means all variables. This name is variable name, without info-
    def getSheet(self, sheet_name, variables=None):
        sheet = self._sheets.get_sheet("info-" + sheet_name)
        if sheet:
            if variables:
                data = [
                    info_node
                    for info_node in sheet.data
                    if info_node.variable in variables
                ]

                sheet.data = data
                sheet.get_sorted_sheet()
                return sheet
            else:
                return sheet
        elif sheet == None:
            return None
        elif len(sheet) == 0:
            return sheet
        else:
            raise ValueError(
                f"Sheet info-{sheet_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models"
            )

    # get sheet object by specific sheet name. This name is variable name, without table-
    def getTable(self, table_name, variables=None):
        table = self._tables.get_table("table-" + table_name)
        if table is not None:
            if variables:
                temp_table = copy.deepcopy(table)
                if len(table.data) > 0:
                    for index, data in enumerate(table.data):
                        for k in data.__dict__:  # check every TableNode object
                            if k not in variables:  # if not in variable list
                                delattr(
                                    temp_table.data[index], k
                                )  # delete from returning obj
                temp_table.column_variables = variables  # update the table variables
                temp_table.column_titles = [
                    table.get_title_by_variable(v) for v in variables
                ]
                # flip the index_variable_pair to get the index by variable
                variable_index_pair = {
                    v: k for k, v in table.index_variable_pair.items()
                }
                temp_table.column_index = [variable_index_pair[v] for v in variables]
                temp_table.get_sorted_table()
                return temp_table
            else:
                return table
        else:
            raise ValueError(
                f"Table table-{table_name} is not existed in {self.excel_name}. Or it was not referrenced in your pydantic models"
            )

    """ check if a row is None. Return True if there is at least one column not None """

    def _notNone(self, row_data):
        return any([False if x == None else True for x in row_data])

    """ Save excel """

    def make_excel(
        self,
        output,
        sheets=None,
        tables=None,
        language=Language.ENGLISH,
        protection=True,
        password="",
    ):
        sheets = sheets if sheets != None else self._sheets._sheets
        tables = tables if tables != None else self._tables._tables
        output = ExcelWritter(
            output_file_name=output,
            sheets=sheets,
            tables=tables,
            language=language,
        )

        output.create(protection=protection, password=password)
